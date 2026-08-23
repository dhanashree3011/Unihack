"""
output_writer.py
------------------
Turns a list of pipeline.RowResult into the final deliverable: an XLSX
(or CSV) with EXACTLY the headers from the expected-output template -
same names, same order, nothing added/removed/renamed, per the brief's
hard requirement.

Headers are read from whatever template file is handed in at runtime
(load_headers()), never hardcoded in this module. That's what lets the
same code run against the eval dataset's own expected-output header set
without a code change, even if a header gets reworded or reordered.

Two write modes:
  write_values_only()   - plain header row + values, fastest, good default
  write_with_confidence_highlighting() - same values, but low-confidence /
      blank cells get a light fill color, so opening the file in Excel
      immediately shows a reviewer where to look. This is the same
      confidence signal the Streamlit review queue uses, just exported
      for anyone who prefers working in a spreadsheet instead.
"""
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from . import pipeline as pipeline_mod
from .pipeline import REVIEW_THRESHOLD

FILL_LOW = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
FILL_BLANK = PatternFill(start_color="F1F1F1", end_color="F1F1F1", fill_type="solid")
FILL_OK = None


def load_headers(expected_output_path: str) -> list:
    if expected_output_path.lower().endswith((".xlsx", ".xls")):
        df = pd.read_excel(expected_output_path, nrows=0)
    else:
        df = pd.read_csv(expected_output_path, nrows=0)
    return [c.strip() for c in df.columns]


def row_to_dict(row_result, headers: list) -> dict:
    """Maps a RowResult onto exactly `headers`, blank string for anything not populated."""
    out = {}
    for h in headers:
        fr = row_result.fields.get(h)
        out[h] = fr.value if fr else ""
    return out


def build_dataframe(row_results: list, headers: list) -> pd.DataFrame:
    rows = [row_to_dict(r, headers) for r in row_results]
    return pd.DataFrame(rows, columns=headers)


def write_csv(row_results: list, headers: list, out_path: str):
    df = build_dataframe(row_results, headers)
    df.to_csv(out_path, index=False)
    return out_path


def write_xlsx_plain(row_results: list, headers: list, out_path: str):
    df = build_dataframe(row_results, headers)
    df.to_excel(out_path, index=False)
    return out_path


def write_xlsx_with_confidence(row_results: list, headers: list, out_path: str,
                                review_threshold: float = REVIEW_THRESHOLD):
    """
    Same values as write_xlsx_plain, plus per-cell fill color so a reviewer
    opening the file directly in Excel (no Streamlit needed) can see at a
    glance which fields to double check: amber = populated but below the
    confidence threshold, grey = left blank because nothing was found.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Enriched Output"

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    for row_idx, r in enumerate(row_results, start=2):
        for col_idx, header in enumerate(headers, start=1):
            fr = r.fields.get(header)
            value = fr.value if fr else ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if not value:
                cell.fill = FILL_BLANK
            elif fr and fr.confidence < review_threshold:
                cell.fill = FILL_LOW

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    wb.save(out_path)
    return out_path


def write_review_log_csv(row_results: list, out_path: str):
    """
    A companion audit file: one row per (row, flagged field) that fell
    below the review threshold, with the source URL/snippet that produced
    it. Useful for judges/reviewers who want to see *why* a field is
    marked low-confidence without opening the Streamlit app.
    """
    records = []
    for r in row_results:
        for field_name in r.review_flags:
            fr = r.fields.get(field_name)
            if not fr:
                continue
            records.append({
                "row_key": r.row_key, "field": field_name, "value": fr.value,
                "confidence": fr.confidence, "source_url": fr.source_url, "snippet": fr.snippet,
            })
    pd.DataFrame(records, columns=["row_key", "field", "value", "confidence", "source_url", "snippet"]) \
        .to_csv(out_path, index=False)
    return out_path
